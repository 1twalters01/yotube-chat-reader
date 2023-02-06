import pytchat
from openpyxl import Workbook, load_workbook


original = input('Click share on the youtube live and then share the link here: ')
link = original.replace('https://www.youtube.com/watch?v=','')
link = link.replace('https://www.youtube.com/live/', '')
link = link.replace('?feature=share','')
chat = pytchat.create(video_id=link.replace('https://www.youtube.com/watch?v=',''))

file_name = input('enter the name of the excel file: ') + ".xlsx"

try:
    workbook = load_workbook(file_name)
    worksheet = workbook.active
except:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet['A1'].value = original
    worksheet['A2'].value = 'Date and time'
    worksheet['B2'].value = 'Username'
    worksheet['C2'].value = 'Message'
    worksheet['D2'].value = 'Amount'
    workbook.save(file_name)

for cell in worksheet["A"]:
    if cell.value is None:
        break
else:
    row = cell.row + 1


while chat.is_alive():
    try:
        for c in chat.get().items:
            if c.type == 'superChat':
                worksheet['A'+str(row)].value = c.datetime
                worksheet['B'+str(row)].value = c.author.name
                worksheet['C'+str(row)].value = c.message
                worksheet['D'+str(row)].value = c.amountString
                workbook.save(file_name)
                row = row+1
    except AttributeError:
        print("Error occurred. Restarting...")
        chat = pytchat.create(video_id=link.replace('https://www.youtube.com/watch?v=',''))