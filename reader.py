import pytchat
from openpyxl import Workbook, load_workbook

link = "https://www.youtube.com/watch?v=MtQFR8xIfVI"
file_name = "example.xlsx"


try:
    workbook = load_workbook(file_name)
    worksheet = workbook.active
except:
    workbook = Workbook()
    worksheet = link
    worksheet['A1'].value = 'date and time'
    worksheet['A2'].value = 'date and time'
    worksheet['B2'].value = 'username'
    worksheet['C2'].value = 'message'
    workbook.save(file_name)


for cell in worksheet["A"]:
    if cell.value is None:
        print(cell.row)
        break
else:
    row = cell.row + 1
print(row)

    

chat = pytchat.create(video_id=link.replace('https://www.youtube.com/watch?v=',''))
while chat.is_alive():
    try:
        for c in chat.get().sync_items():
            if c.type == 'superChat':
                print(f"{c.datetime} [{c.author.name}] - {c.message}")
                worksheet['A'+str(row)].value = c.datetime
                worksheet['B'+str(row)].value = c.author.name
                worksheet['C'+str(row)].value = c.message
                workbook.save(file_name)
                row = row+1
    except AttributeError:
        print("Error occurred. Restarting...")
        chat = pytchat.create(video_id="uIx8l2xlYVY")