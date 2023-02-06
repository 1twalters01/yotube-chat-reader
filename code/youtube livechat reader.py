import pytchat
from openpyxl import Workbook, load_workbook
import os
from pathlib import Path

original = input('Click share on the youtube live and then paste the link here: ')
link = original.replace('https://www.youtube.com/watch?v=','')
link = link.replace('https://www.youtube.com/live/', '')
link = link.replace('?feature=share','')

file_name = input('Enter the name of the excel file without a file extension: ') + ".xlsx"
location = input('Enter the file location or leave blank to place inside this directory: ')

if location == '':
    current = Path(os.getcwd())
    location = str(current.parents[0])+'\\'+file_name

try:
    workbook = load_workbook(location)
    worksheet = workbook.active
except:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet['A1'].value = original
    worksheet['A2'].value = 'Date and time'
    worksheet['B2'].value = 'Username'
    worksheet['C2'].value = 'Message'
    worksheet['D2'].value = 'Amount'
    workbook.save(location)

for cell in worksheet["A"]:
    if cell.value is None:
        break
else:
    row = cell.row + 1

chat = pytchat.create(video_id=link)
while chat.is_alive():
    try:
        for c in chat.get().sync_items():
            if c.type == 'superChat':
                print(c.datetime, '[', c.author.name, '] - ', c.message, '-', c.amountString)
                worksheet['A'+str(row)].value = c.datetime
                worksheet['B'+str(row)].value = c.author.name
                worksheet['C'+str(row)].value = c.message
                worksheet['D'+str(row)].value = c.amountString
                workbook.save(location)
                row = row+1
    except AttributeError:
        print("Error occurred. Restarting...")
        chat = pytchat.create(video_id=link.replace('https://www.youtube.com/watch?v=',''))